"""Aggregator agent — generates CSV and Excel output files."""

from __future__ import annotations

import time
from pathlib import Path

import structlog

from job_hunter_agents.agents.base import BaseAgent
from job_hunter_core.state import PipelineState

logger = structlog.get_logger()


class AggregatorAgent(BaseAgent):
    """Generate output files (CSV, Excel) from scored jobs."""

    agent_name = "aggregator"

    async def run(self, state: PipelineState) -> PipelineState:
        """Write scored jobs to output files."""
        self._log_start({"scored_jobs_count": len(state.scored_jobs)})
        start = time.monotonic()

        output_dir = self.settings.output_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        rows = self._build_rows(state)
        output_files: list[str] = []

        if "csv" in state.config.output_formats:
            csv_path = output_dir / f"{state.config.run_id}_results.csv"
            self._write_csv(rows, csv_path)
            output_files.append(str(csv_path))

        if "xlsx" in state.config.output_formats:
            xlsx_path = output_dir / f"{state.config.run_id}_results.xlsx"
            self._write_excel(rows, xlsx_path, state)
            output_files.append(str(xlsx_path))

        state.run_result = state.build_result(
            status="success" if state.scored_jobs else "partial",
            duration_seconds=time.monotonic() - start,
            output_files=output_files,
        )

        self._log_end(
            time.monotonic() - start,
            {
                "output_files": output_files,
            },
        )
        return state

    def _build_rows(self, state: PipelineState) -> list[dict[str, object]]:
        """Build output rows from scored jobs (best job per company)."""
        # Build company_id -> tier lookup
        tier_map: dict[str, str] = {}
        for company in state.companies:
            tier_map[str(company.id)] = company.tier.value

        # Deduplicate: keep only the highest-scored job per company
        # (scored_jobs are already sorted by score descending)
        seen_companies: set[str] = set()
        deduped = []
        for sj in state.scored_jobs:
            co_name = sj.job.company_name
            if co_name in seen_companies:
                continue
            seen_companies.add(co_name)
            deduped.append(sj)

        rows: list[dict[str, object]] = []
        for rank, sj in enumerate(deduped, start=1):
            job = sj.job
            report = sj.fit_report
            salary = self._format_salary(job.salary_min, job.salary_max, job.currency)
            tier = tier_map.get(str(job.company_id), "unknown")

            rows.append(
                {
                    "Rank": rank,
                    "Score": report.score,
                    "Recommendation": report.recommendation,
                    "Company": job.company_name,
                    "Company Tier": tier,
                    "Title": job.title,
                    "Location": job.location or "",
                    "Remote Type": job.remote_type,
                    "Posted Date": str(job.posted_date) if job.posted_date else "",
                    "Salary Range": salary,
                    "Skill Match": ", ".join(report.skill_overlap),
                    "Skill Gaps": ", ".join(report.skill_gaps),
                    "Fit Summary": report.summary,
                    "Apply URL": str(job.apply_url),
                }
            )
        return rows

    @staticmethod
    def _format_salary(
        salary_min: int | None,
        salary_max: int | None,
        currency: str | None,
    ) -> str:
        """Format salary range with appropriate currency symbol."""
        symbols: dict[str, str] = {
            "USD": "$",
            "INR": "₹",
            "EUR": "€",
            "GBP": "£",
            "CAD": "C$",
            "AUD": "A$",
            "SGD": "S$",
        }
        cur = (currency or "USD").upper()
        sym = symbols.get(cur, f"{cur} ")

        if salary_min and salary_max:
            return f"{sym}{salary_min:,}-{sym}{salary_max:,} {cur}"
        if salary_min:
            return f"{sym}{salary_min:,}+ {cur}"
        return ""

    def _write_csv(self, rows: list[dict[str, object]], path: Path) -> None:
        """Write rows to CSV file."""
        import csv

        if not rows:
            return

        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

        logger.info("csv_written", path=str(path), rows=len(rows))

    def _write_excel(
        self,
        rows: list[dict[str, object]],
        path: Path,
        state: PipelineState,
    ) -> None:
        """Write rows to Excel with formatting."""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        if not rows:
            return

        df = pd.DataFrame(rows)
        df.to_excel(str(path), index=False, sheet_name="Results")

        wb = load_workbook(str(path))
        ws = wb["Results"]

        # Conditional formatting for score column
        green = PatternFill(start_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", fill_type="solid")

        score_col = 2  # Column B
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=score_col)
            if isinstance(cell.value, int):
                if cell.value >= 80:
                    cell.fill = green
                elif cell.value >= 60:
                    cell.fill = yellow

        # Make Apply URL column hyperlinked (14th column with Company Tier added)
        url_col = 14  # Column N
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=url_col)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="0563C1", underline="single")

        # Run summary sheet
        summary_ws = wb.create_sheet("Run Summary")
        summary_data = [
            ("Run ID", state.config.run_id),
            ("Companies Attempted", len(state.companies)),
            ("Jobs Scraped", len(state.raw_jobs)),
            ("Jobs Scored", len(state.scored_jobs)),
            ("Total Tokens", state.total_tokens),
            ("Estimated Cost (USD)", f"${state.total_cost_usd:.2f}"),
            ("Errors", len(state.errors)),
        ]
        for i, (key, value) in enumerate(summary_data, start=1):
            summary_ws.cell(row=i, column=1, value=key)
            summary_ws.cell(row=i, column=2, value=str(value))

        wb.save(str(path))
        logger.info("excel_written", path=str(path), rows=len(rows))
